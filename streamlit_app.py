
import streamlit as st
import pandas as pd
from io import BytesIO
import os
from PIL import Image

st.set_page_config(page_title="LYN – Agente de Cruce", layout="centered")

# Mostrar logotipo si existe
logo_path = os.path.join("assets", "logo.png")
try:
    st.image(Image.open(logo_path), width=150)
except:
    st.warning("⚠️ No se pudo cargar el logo.")
    st.image(Image.open(logo_path), width=150)

st.title("LYN DE MEXICO – Agente de Cruce de Archivos")
st.markdown("Sube los archivos requeridos para generar el reporte final consolidado:")

archivo_skus = st.file_uploader("1. Cargar archivo SKUS", type="xlsx")
archivo_existencias = st.file_uploader("2. Cargar archivo EXISTENCIAS", type="xlsx")
archivo_ordenado = st.file_uploader("3. Cargar archivo ORDENADO", type="xlsx")
archivo_liverpool = st.file_uploader("4. Cargar archivo VENTAS LIVERPOOL", type="xlsx")
archivo_palacio = st.file_uploader("5. Cargar archivo VENTAS PALACIO", type="xlsx")

if st.button("🚀 Ejecutar Cruce"):
    if not all([archivo_skus, archivo_existencias, archivo_ordenado, archivo_liverpool, archivo_palacio]):
        st.warning("Por favor carga todos los archivos.")
    else:
        df_skus = pd.read_excel(archivo_skus)
        df_skus = df_skus[df_skus['CODIGO'].notna()].copy()
        df_skus['CODIGO'] = df_skus['CODIGO'].astype(str).str.strip()

        df_existencias = pd.read_excel(archivo_existencias, header=3)
        df_existencias = df_existencias[['CODIGO', 'CANTIDAD']].dropna()
        df_existencias['CODIGO'] = df_existencias['CODIGO'].astype(str).str.strip()
        df_existencias.rename(columns={'CANTIDAD': 'EXISTENCIAS'}, inplace=True)

        df_ordenado = pd.read_excel(archivo_ordenado, header=4)
        df_ordenado = df_ordenado[['CODIGO', 'PEDIDO']].dropna()
        df_ordenado['CODIGO'] = df_ordenado['CODIGO'].astype(str).str.strip()
        df_ordenado.rename(columns={'PEDIDO': 'ORDENADO'}, inplace=True)

        # Unir por CODIGO
        df_skus = df_skus.merge(df_existencias, how='left', on='CODIGO')
        df_skus = df_skus.merge(df_ordenado, how='left', on='CODIGO')

        # VENTAS LIVERPOOL
        df_ventas_liv = pd.read_excel(archivo_liverpool)
        df_ventas_liv = df_ventas_liv[['Artículo', 'vta total 9 meses']].dropna()
        df_ventas_liv.columns = ['CODIGO_LIV', 'VENTAS LIV 9 MESES']
        df_ventas_liv['CODIGO_LIV'] = df_ventas_liv['CODIGO_LIV'].astype(float).astype('Int64').astype(str)
        df_skus['LIVERPOOL '] = df_skus['LIVERPOOL '].astype(float).astype('Int64').astype(str)
        df_skus = pd.merge(df_skus, df_ventas_liv, how='left', left_on='LIVERPOOL ', right_on='CODIGO_LIV')
        df_skus.drop(columns='CODIGO_LIV', inplace=True)

        # VENTAS PALACIO
        df_ventas_pal = pd.read_excel(archivo_palacio)
        df_ventas_pal = df_ventas_pal[['Clave de Artículo', 'Venta Neta en UM']].dropna()
        df_ventas_pal.columns = ['CODIGO_PAL', 'VENTAS PAL 9 MESES']
        df_ventas_pal['CODIGO_PAL'] = df_ventas_pal['CODIGO_PAL'].astype(str).str.strip()
        df_ventas_pal = df_ventas_pal[df_ventas_pal['CODIGO_PAL'].str.isnumeric()]
        df_ventas_pal['CODIGO_PAL'] = df_ventas_pal['CODIGO_PAL'].astype(float).astype('Int64').astype(str)
        df_skus['PALACIO'] = df_skus['PALACIO'].astype(float).astype('Int64').astype(str)
        df_skus = pd.merge(df_skus, df_ventas_pal, how='left', left_on='PALACIO', right_on='CODIGO_PAL')
        df_skus.drop(columns='CODIGO_PAL', inplace=True)

        # Vista previa específica
        st.subheader("✅ Vista previa: Códigos con EXISTENCIAS y ORDENADO")
        st.dataframe(df_skus[['CODIGO', 'EXISTENCIAS', 'ORDENADO']].head(20))

        # Ordenar columnas en orden deseado
        columnas_ordenadas = [
            "CODIGO", "DESCRIPCION", "LIVERPOOL ", "PALACIO",
            "ORDENADO", "EXISTENCIAS", "VENTAS LIV 9 MESES", "VENTAS PAL 9 MESES"
        ]
        columnas_presentes = [col for col in columnas_ordenadas if col in df_skus.columns]
        df_skus = df_skus[columnas_presentes + [col for col in df_skus.columns if col not in columnas_presentes]]
    
        cols = df_skus.columns.tolist()
        for col in ['EXISTENCIAS', 'ORDENADO']:
            if col in cols:
                cols.remove(col)
                cols.insert(cols.index('CODIGO') + 1, col)
        df_skus = df_skus[cols]

        
        # Exportar archivo con formato condicional usando openpyxl
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name

        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            df_skus.to_excel(writer, index=False, sheet_name="Reporte")
        
        wb = load_workbook(temp_path)
        ws = wb.active

        rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            col_index = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "EXISTENCIAS":
                    col_index = idx
                    break
            if col_index:
                celda = row[col_index - 1]
                if celda.value is None or celda.value == 0:
                    for c in row:
                        c.fill = rojo

        wb.save(temp_path)

        with open(temp_path, "rb") as f:
            st.download_button("📥 Descargar archivo final", data=f, file_name="Reporte_LYN_Final.xlsx")
    
        buffer = BytesIO()
        df_skus.to_excel(buffer, index=False)
        buffer.seek(0)

        st.success("🎉 Cruce completo. Puedes descargar tu archivo.")
        # Botón eliminado por duplicado
